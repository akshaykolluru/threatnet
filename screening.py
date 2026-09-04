from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree
try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # Optional for API-only deployments.
    load_workbook = None


def _read_minimal_xlsx(path: Path) -> list[dict]:
    """Read the flat tabular workbook shape used by the demo without extra packages."""
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with ZipFile(path) as archive:
        shared = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            shared = ["".join(node.itertext()) for node in root.findall(f"{ns}si")]
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        first_sheet = workbook.find(f"{ns}sheets/{ns}sheet")
        rel_id = first_sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target = next(node.attrib["Target"] for node in rels if node.attrib.get("Id") == rel_id)
        sheet_path = target.lstrip("/")
        if not sheet_path.startswith("xl/"):
            sheet_path = "xl/" + sheet_path
        root = ElementTree.fromstring(archive.read(sheet_path))
        output = []
        for row in root.findall(f".//{ns}sheetData/{ns}row"):
            values = []
            for cell in row.findall(ns + "c"):
                value = cell.find(ns + "v")
                text = "" if value is None else value.text or ""
                if cell.attrib.get("t") == "s" and text:
                    text = shared[int(text)]
                values.append(text)
            output.append(values)
    headers, data = output[0], output[1:]
    return [dict(zip(headers, row)) for row in data if any(row)]


def screen_workbook(path: Path, limit: int = 50) -> dict:
    if load_workbook is None:
        records = _read_minimal_xlsx(path)
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        suspect_sheet = workbook["Suspects"] if "Suspects" in workbook.sheetnames else workbook.active
        rows = suspect_sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        records = [dict(zip(headers, row)) for row in rows if any(value is not None for value in row)]
    scored = []
    for record in records:
        route = str(record.get("Route Match", "")).lower()
        calls = float(record.get("Call Link Count", 0) or 0)
        score = min(1.0, (0.48 if route == "high" else 0.25 if route == "medium" else 0.08) + min(calls, 8) * 0.05)
        scored.append({"suspect_id": record.get("Suspect ID", ""), "name": record.get("Name", ""), "area": record.get("Area", ""), "vehicle": record.get("Vehicle", ""), "route_match": record.get("Route Match", ""), "call_link_count": int(calls), "score": round(score, 4), "reasons": [f"route match: {record.get('Route Match', 'unknown')}", f"linked calls: {int(calls)}"]})
    scored.sort(key=lambda item: item["score"], reverse=True)
    return {"input_rows": len(records), "shortlist_size": min(limit, len(scored)), "shortlist": scored[:limit], "interpretation": "Ranked review queue based on declared spreadsheet signals. This does not identify a person or establish guilt."}
